import requests
import random
import pytz
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook
from Insults import insult_firstword, insult_secondword, insult_thirdword

eltp_tz = pytz.timezone('Europe/London')  # Timezone that game times are given in are in Lodon
ownteam = 'TC Ballieveldert'
date_format = '%d/%m'
time_format = '%A:%H:%M'

rawdata_headers = [
    'date', 'time', 'type', 'week', 'map1_name', 'map1_link', 'map2_name', 'map2_link', 'map3_name', 'map3_link',
    'own_team', 'opponent'
]

matchday_msg_timeoffset_mins_list = [8 * 60, (1 * 60) + 15, 5]  # timeoffset in minutes
scrim_msg_timeoffset_mins_list = [8 * 60, 5]  # timeoffset in minutes

scrim_offset = -5


def is_dst(tz):
    now = pytz.utc.localize(datetime.utcnow())
    return now.astimezone(tz).dst() != timedelta(0)


def conv_eventinfo_to_dict(datetime, type):
    return {
        'datetime': datetime,
        'type': type,
    }


def conv_dt_to_time(a_datetime):
    return datetime.strftime(a_datetime, time_format)


def conv_dt_to_date(a_datetime):
    return datetime.strftime(a_datetime, date_format)


def update_fixtures_csv():
    minorsfixtures_url = 'https://docs.google.com/spreadsheets/d/1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM/export?format=xlsx&id=1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM&gid=92829842'
    req_content = requests.get(url=minorsfixtures_url).content
    with open('data/minorsfixtures_tmp.xlsx', 'wb') as f:
        f.write(req_content)
    ws_curr = load_workbook(filename='data/minorsfixtures.xlsx', data_only=False)['Minors Fixtures']
    ws_tmp = load_workbook(filename='data/minorsfixtures_tmp.xlsx', data_only=False)['Minors Fixtures']
    for row_idx, row in enumerate(ws_curr.rows):
        for col_idx, cell in enumerate(row):
            #  only update fixture csv file if excel file has changed, this allows manual changes to csv to persist longer
            if ws_curr.cell(row_idx+1, col_idx+1).value != ws_tmp.cell(row_idx+1, col_idx+1).value:
                convert_sheet_to_rawcsv()



def convert_sheet_to_rawcsv():
    ws_fixt = load_workbook(filename='data/minorsfixtures.xlsx', data_only=False)['Minors Fixtures']
    ws_fixt_dataonly = load_workbook(filename='data/minorsfixtures.xlsx', data_only=True)['Minors Fixtures']

    with open('data/rawfixtures.csv', 'w') as f:
        teamcol1 = 1  # first column with team name on https://docs.google.com/spreadsheets/d/1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM/edit#gid=92829842
        teamcol2 = 7  # second column with team name on https://docs.google.com/spreadsheets/d/1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM/edit#gid=92829842
        for row_idx, row in enumerate(ws_fixt_dataonly.rows):  # go through all the rows of the fixtures tab
            daterowinfo_list = []
            opponent = ''
            if isinstance(datetime_sheet := row[0].value,
                          datetime):  # find the row that matches the current date, e.g. day on which map is played
                for col_idx, curr_cell in enumerate(row):
                    daterowinfo_list.append((curr_cell.value, ws_fixt.cell(row=row_idx + 1,
                                                                           column=col_idx + 1).value))  # adding cell value from curr_cell.value and formula from ws_fixt.cell to get link to map
                for row_opp_idx in range(row_idx + 2,
                                         row_idx + 6):  # from the row in which the date of the week and maps are specified, the actual matches are gives in this range of rows
                    if ws_fixt.cell(row=row_opp_idx + 1,
                                    column=teamcol1).value == ownteam:  # if we are team 1 then our opponent for this week is team 2
                        opponent = ws_fixt.cell(row=row_opp_idx + 1, column=teamcol2).value.strip()
                        break
                    elif ws_fixt.cell(row=row_opp_idx + 1, column=teamcol2).value == ownteam:  # and vice versa
                        opponent = ws_fixt.cell(row=row_opp_idx + 1, column=teamcol1).value.strip()
                        break

                week, _ = daterowinfo_list[8]

                process_map_link = lambda x: x.split('"')[1] if len(
                    (x := '???' if x is None else x).split("'")) > 1 else x
                process_map_name = lambda x: '???' if x is None else x

                map1_name, map1_link = daterowinfo_list[1]
                map1_link = process_map_link(map1_link)
                map1_name = process_map_name(map1_name)
                map2_name, map2_link = daterowinfo_list[14]
                map2_link = process_map_link(map2_link)
                map2_name = process_map_name(map2_name)
                map3_name, map3_link = daterowinfo_list[22]
                map3_link = process_map_link(map3_link)
                map3_name = process_map_name(map3_name)

                match_datetime = datetime(year=datetime_sheet.year, month=datetime_sheet.month, day=datetime_sheet.day,
                                          hour=(19 if is_dst(eltp_tz) else 20), minute=0,
                                          tzinfo=pytz.utc)  # for calculations it is recommended to store times in utc, so we have to correct for that wrt actual game times
                scrim_datetime = match_datetime + timedelta(
                    days=scrim_offset)  # scrims 5 days before actual game e.g. Wednesdays
                base_data = [week, map1_name, map1_link, map2_name, map2_link, map3_name, map3_link, ownteam, opponent]


                match_data = [conv_dt_to_date(match_datetime), conv_dt_to_time(match_datetime), 'match'] + base_data
                print(','.join(match_data), file=f)  # NOT A LOG, actually writing!!
                scrim_data = [conv_dt_to_date(scrim_datetime), conv_dt_to_time(scrim_datetime), 'scrim'] + base_data
                print(','.join(scrim_data), file=f)  # NOT A LOG, actually writing!!


def update_table_csv():
    minorstable_url = 'https://docs.google.com/spreadsheets/d/1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM/export?format=xlsx&id=1ruIpfqYwHyH17tvN6PVEYqYJm8geUoqgdsGgzm-BjhM&gid=601872256'
    open('data/minorstable.xlsx', 'wb').write(requests.get(url=minorstable_url).content)


def get_ranking(team):
    ws_table_dataonly = load_workbook(filename='data/minorstable.xlsx', data_only=True)['Minors Table']
    for row in ws_table_dataonly.iter_rows():
        if team == row[2].value:
            return row[0].value


def ranking_to_placement(ranking):
    if ranking == 1:
        return f'{str(ranking)}st'
    elif ranking == 2:
        return f'{str(ranking)}nd'
    else:
        return f'{str(ranking)}th'


def create_msg(a_event):
    own_ranking = ranking_to_placement(get_ranking(a_event['own_team']))
    opponent_ranking = ranking_to_placement(get_ranking(a_event['opponent']))

    current_time = conv_dt_to_time(datetime.utcnow())
    current_time_mins = int(current_time.split(':')[1]) * 60 + int(
        current_time.split(':')[2])  # time_format = '%A:%H:%M'

    important_offset_mins = matchday_msg_timeoffset_mins_list[0]

    msg = ''
    for timeoffset_mins in matchday_msg_timeoffset_mins_list:
        event_msg_time_mins = int(a_event['time'].split(':')[1]) * 60 + int(
            a_event['time'].split(':')[2]) - timeoffset_mins
        if current_time_mins == event_msg_time_mins:
            hours_val = timeoffset_mins // 60
            mins_val = timeoffset_mins % 60

            msg = ''
            if a_event['type'] == 'match':
                if timeoffset_mins == important_offset_mins:
                    msg = f"@everyone MATCHDAY - Today we're playing {a_event['opponent']} (ranked {opponent_ranking}) on {a_event['map1_name']}, {a_event['map2_name']} & {a_event['map3_name']} at 8pm BST" \
                          f" that's in less than {hours_val} hours and {mins_val} minutes." \
                          f" Lets beat those {random.choice(insult_firstword)} {random.choice(insult_secondword)} {random.choice(insult_thirdword)}!!" \
                          f" Please be on at 7pm BST for scrims." \
                          f"\nWe're currently in {a_event['week'].lower()} out of 7 weeks of the regular season and are ranked {own_ranking}."
                else:
                    msg = f" MATCHDAY REMINDER - Today we're playing {a_event['opponent']} (ranked {opponent_ranking}) on {a_event['map1_name']}, {a_event['map2_name']} & {a_event['map3_name']} at 8pm BST" \
                          f" that's in less than {hours_val} hours and {mins_val} minutes." \
                          f" Please be on at 7pm BST for scrims."
            elif a_event['type'] == 'scrim':
                msg = f"PRACTICE - Playing {a_event['opponent']} on {a_event['map1_name']}, {a_event['map2_name']} & {a_event['map3_name']} this week" \
                      f" Please be on at 8pm BST for practice." \
                      f" that's in less than {hours_val} hours and {mins_val} minutes."
            if len(msg) > 0 and (a_event['type'] == 'scrim' or timeoffset_mins == important_offset_mins):
                msg += f"\n{a_event['map1_name']}: {a_event['map1_link']}," \
                       f" {a_event['map2_name']}: {a_event['map2_link']}," \
                       f" {a_event['map3_name']}: {a_event['map3_link']}"
            break
    return msg


def get_message():
    with open('data/rawfixtures.csv', 'r') as f:
        for line in f:
            line_list = line.strip().split(',')
            if conv_dt_to_date(datetime.utcnow()) == line_list[0]:
                a_event = dict(zip(rawdata_headers, line_list))
                return create_msg(a_event)
    return ''
