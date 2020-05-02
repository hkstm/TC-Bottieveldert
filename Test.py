import requests
from datetime import datetime, timedelta
import csv
import json
from openpyxl import load_workbook
import pytz

ownteam = 'TC Ballieveldert'
ws_table_dataonly = load_workbook(filename='data/minorstable.xlsx', data_only=True)['Minors Table']
for row in ws_table_dataonly.iter_rows():
    if ownteam == row[2].value:
        print(row[0].value)